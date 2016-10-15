#-*- coding: UTF-8 -*-
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
from openpyxl import load_workbook

def Create_User(vorname, nachname, kurs):
    user = User.objects.create_user('{}.{}'.format(vorname,nachname), password='{}.{}'.format(nachname,vorname))
    user.is_superuser = False
    user.is_staff = False
    user.save()
    g = Group.objects.get(name=kurs)
    g.user_set.add(user)

def Read_Excel(xls_file):
    wb = load_workbook(filename=xls_file)
    sheet = wb.get_sheet_by_name('Kontaktinfos')
    for i in range(3,sheet.max_row):
        vorname.append(sheet.cell(row=i, column=1).value.lower().replace(" ", "_").strip('\n'))
        nachname.append(sheet.cell(row=i, column=2).value.lower().replace(" ", "_").strip('\n'))
    #print(vorname, nachname)
        #   print(sheet_ranges['A3'].value)


vorname = []
nachname = []
Read_Excel('Physik8b.xlsx')
for i in range(0, len(vorname)):
    Create_User(vorname[i], nachname[i], '8bPH1617')